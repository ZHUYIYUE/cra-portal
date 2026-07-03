#!/usr/bin/env python3
"""
CRA Portal - 临床研究助理管理中心
后端 v2.0 - Supabase 版本
数据库从 Render PostgreSQL 迁移到 Supabase
"""

import os
import json
import uuid
import io
from datetime import datetime, date
from flask import Flask, render_template, jsonify, request, send_file, Response
from flask_cors import CORS
from supabase import create_client, Client

app = Flask(__name__)
CORS(app)

# ========== Supabase 配置 ==========
SUPABASE_URL = os.environ.get("SUPABASE_URL", "")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY", "")

if not SUPABASE_URL or not SUPABASE_KEY:
    print("=" * 50)
    print("⚠️  警告: SUPABASE_URL 和 SUPABASE_KEY 环境变量未设置！")
    print("请在 Render Dashboard 设置环境变量：")
    print("  SUPABASE_URL  = https://xxx.supabase.co")
    print("  SUPABASE_KEY  = eyJxxx...")
    print("=" * 50)

supabase: Client = None
if SUPABASE_URL and SUPABASE_KEY:
    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)


def get_supabase():
    """获取 Supabase 客户端"""
    if not supabase:
        raise RuntimeError("Supabase 未配置，请设置环境变量")
    return supabase


def gen_id():
    """生成8位短ID"""
    return uuid.uuid4().hex[:8]


def parse_date(date_str):
    """安全解析日期字符串"""
    if not date_str:
        return None
    try:
        return datetime.fromisoformat(date_str).strftime("%Y-%m-%dT%H:%M:%S")
    except (ValueError, TypeError):
        try:
            return datetime.strptime(date_str, "%Y-%m-%d").strftime("%Y-%m-%dT%H:%M:%S")
        except (ValueError, TypeError):
            return date_str


# ========== 页面路由 ==========

@app.route('/')
def index():
    return render_template('index.html')


# ========== 项目 API ==========

@app.route('/api/projects')
def get_projects():
    """获取所有项目列表"""
    try:
        sb = get_supabase()
        result = sb.table('projects').select('*').order('created_at', desc=False).execute()
        projects = []
        for row in result.data:
            # 查询关联统计
            task_count = sb.table('tasks').select('id', count='exact').eq('project_id', row['id']).execute().count
            done_count = sb.table('tasks').select('id', count='exact').eq('project_id', row['id']).eq('done', True).execute().count
            center_count = sb.table('centers').select('id', count='exact').eq('project_id', row['id']).execute().count
            projects.append({
                "id": row['id'],
                "name": row['name'],
                "code": row.get('code', ''),
                "stage": row.get('stage', ''),
                "center_count": center_count,
                "task_count": task_count,
                "done_count": done_count,
                "dbl_date": row.get('dbl_date', ''),
                "notes": row.get('notes', ''),
                "created_at": row.get('created_at', ''),
                "updated_at": row.get('updated_at', '')
            })
        return jsonify({"success": True, "projects": projects})
    except Exception as e:
        print(f"获取项目列表失败: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/project/<project_id>')
def get_project(project_id):
    """获取单个项目详情"""
    try:
        sb = get_supabase()
        result = sb.table('projects').select('*').eq('id', project_id).execute()
        if not result.data:
            return jsonify({"success": False, "error": "项目不存在"}), 404

        row = result.data[0]
        return jsonify({
            "success": True,
            "id": row['id'],
            "name": row['name'],
            "code": row.get('code', ''),
            "stage": row.get('stage', ''),
            "dbl_date": row.get('dbl_date', ''),
            "notes": row.get('notes', ''),
            "created_at": row.get('created_at', ''),
            "updated_at": row.get('updated_at', '')
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/project', methods=['POST'])
def create_project():
    """创建新项目"""
    try:
        sb = get_supabase()
        data = request.json
        new_id = gen_id()

        sb.table('projects').insert({
            "id": new_id,
            "name": data.get('name', ''),
            "code": data.get('code', ''),
            "stage": data.get('stage', ''),
            "dbl_date": data.get('dbl_date', ''),
            "notes": data.get('notes', ''),
            "center_count": 0,
            "created_at": datetime.now().isoformat(),
            "updated_at": datetime.now().isoformat()
        }).execute()

        return jsonify({"success": True, "id": new_id})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/project/<project_id>', methods=['PUT'])
def update_project(project_id):
    """更新项目"""
    try:
        sb = get_supabase()
        data = request.json
        update_data = {
            "updated_at": datetime.now().isoformat()
        }
        for field in ['name', 'code', 'stage', 'dbl_date', 'notes']:
            if field in data:
                update_data[field] = data[field]

        sb.table('projects').update(update_data).eq('id', project_id).execute()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/project/<project_id>', methods=['DELETE'])
def delete_project(project_id):
    """删除项目（级联删除关联数据）"""
    try:
        sb = get_supabase()
        sb.table('projects').delete().eq('id', project_id).execute()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# ========== 中心 API ==========

@app.route('/api/centers')
def get_centers():
    """获取中心列表"""
    try:
        sb = get_supabase()
        project_id = request.args.get('project_id', '')

        query = sb.table('centers').select('*')
        if project_id:
            query = query.eq('project_id', project_id)

        result = query.order('created_at', desc=False).execute()
        centers = []
        for row in result.data:
            task_count = sb.table('tasks').select('id', count='exact').eq('center_id', row['id']).execute().count
            finding_count = sb.table('findings').select('id', count='exact').eq('center_id', row['id']).execute().count
            open_finding_count = sb.table('findings').select('id', count='exact').eq('center_id', row['id']).eq('status', 'Open').execute().count
            centers.append({
                "id": row['id'],
                "project_id": row['project_id'],
                "code": row.get('code', ''),
                "name": row['name'],
                "department": row.get('department', ''),
                "address": row.get('address', ''),
                "pi_name": row.get('pi_name', ''),
                "pi_phone": row.get('pi_phone', ''),
                "pi_email": row.get('pi_email', ''),
                "contact_crc": row.get('contact_crc', ''),
                "contact_crc_phone": row.get('contact_crc_phone', ''),
                "contact_ethics": row.get('contact_ethics', ''),
                "milestones": row.get('milestones', {}),
                "notes": row.get('notes', ''),
                "task_count": task_count,
                "finding_count": finding_count,
                "open_finding_count": open_finding_count,
                "created_at": row.get('created_at', ''),
                "updated_at": row.get('updated_at', '')
            })
        return jsonify({"success": True, "centers": centers})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/center/<center_id>')
def get_center_detail(center_id):
    """获取中心详情"""
    try:
        sb = get_supabase()
        result = sb.table('centers').select('*').eq('id', center_id).execute()
        if not result.data:
            return jsonify({"success": False, "error": "中心不存在"}), 404

        row = result.data[0]

        # 获取关联数据
        staff = sb.table('research_staff').select('*').eq('center_id', center_id).execute().data
        ethics = sb.table('ethics_submissions').select('*').eq('center_id', center_id).execute().data
        deviations = sb.table('protocol_deviations').select('*').eq('center_id', center_id).execute().data
        tasks = sb.table('tasks').select('*').eq('center_id', center_id).execute().data
        findings = sb.table('findings').select('*').eq('center_id', center_id).execute().data

        return jsonify({
            "success": True,
            "id": row['id'],
            "project_id": row['project_id'],
            "code": row.get('code', ''),
            "name": row['name'],
            "department": row.get('department', ''),
            "address": row.get('address', ''),
            "pi_name": row.get('pi_name', ''),
            "pi_phone": row.get('pi_phone', ''),
            "pi_email": row.get('pi_email', ''),
            "contact_crc": row.get('contact_crc', ''),
            "contact_crc_phone": row.get('contact_crc_phone', ''),
            "contact_ethics": row.get('contact_ethics', ''),
            "milestones": row.get('milestones', {}),
            "notes": row.get('notes', ''),
            "staff": staff,
            "ethics_submissions": ethics,
            "protocol_deviations": deviations,
            "tasks": tasks,
            "findings": findings,
            "created_at": row.get('created_at', ''),
            "updated_at": row.get('updated_at', '')
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/center', methods=['POST'])
def create_center():
    """创建中心"""
    try:
        sb = get_supabase()
        data = request.json
        new_id = gen_id()

        sb.table('centers').insert({
            "id": new_id,
            "project_id": data.get('project_id', ''),
            "code": data.get('code', ''),
            "name": data.get('name', ''),
            "department": data.get('department', ''),
            "address": data.get('address', ''),
            "pi_name": data.get('pi_name', ''),
            "pi_phone": data.get('pi_phone', ''),
            "pi_email": data.get('pi_email', ''),
            "contact_crc": data.get('contact_crc', ''),
            "contact_crc_phone": data.get('contact_crc_phone', ''),
            "contact_ethics": data.get('contact_ethics', ''),
            "milestones": data.get('milestones', {}),
            "notes": data.get('notes', ''),
            "created_at": datetime.now().isoformat(),
            "updated_at": datetime.now().isoformat()
        }).execute()

        return jsonify({"success": True, "id": new_id})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/center/<center_id>', methods=['PUT'])
def update_center(center_id):
    """更新中心"""
    try:
        sb = get_supabase()
        data = request.json
        update_data = {"updated_at": datetime.now().isoformat()}
        for field in ['code', 'name', 'department', 'address', 'pi_name', 'pi_phone',
                       'pi_email', 'contact_crc', 'contact_crc_phone', 'contact_ethics',
                       'milestones', 'notes']:
            if field in data:
                update_data[field] = data[field]

        sb.table('centers').update(update_data).eq('id', center_id).execute()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/center/<center_id>', methods=['DELETE'])
def delete_center(center_id):
    """删除中心"""
    try:
        sb = get_supabase()
        sb.table('centers').delete().eq('id', center_id).execute()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# ========== 任务 API ==========

@app.route('/api/tasks')
def get_tasks():
    """获取任务列表"""
    try:
        sb = get_supabase()
        project_id = request.args.get('project_id', '')

        query = sb.table('tasks').select('*')
        if project_id:
            query = query.eq('project_id', project_id)

        result = query.order('created_at', desc=False).execute()
        tasks = []
        for row in result.data:
            # 获取中心名称
            center_name = ""
            if row.get('center_id'):
                c_result = sb.table('centers').select('name', 'code').eq('id', row['center_id']).execute()
                if c_result.data:
                    c = c_result.data[0]
                    center_name = f"{c.get('code', '')} {c.get('name', '')}".strip()

            # 获取项目名称
            p_result = sb.table('projects').select('name').eq('id', row['project_id']).execute()
            project_name = p_result.data[0]['name'] if p_result.data else ""

            tasks.append({
                "id": row['id'],
                "title": row['title'],
                "project_id": row['project_id'],
                "project_name": project_name,
                "center_id": row.get('center_id', ''),
                "center_name": center_name,
                "priority": row.get('priority', 'medium'),
                "ability_type": row.get('ability_type', 'execution'),
                "due_date": row.get('due_date', ''),
                "done": row.get('done', False),
                "task_status": row.get('task_status', 'pending'),
                "created_at": row.get('created_at', '')
            })
        return jsonify({"success": True, "tasks": tasks})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/task', methods=['POST'])
def create_task():
    """创建任务"""
    try:
        sb = get_supabase()
        data = request.json
        new_id = data.get('id', gen_id())

        sb.table('tasks').insert({
            "id": new_id,
            "title": data.get('title', ''),
            "project_id": data.get('project_id', ''),
            "center_id": data.get('center_id', ''),
            "priority": data.get('priority', 'medium'),
            "ability_type": data.get('ability_type', 'execution'),
            "due_date": data.get('due_date', ''),
            "done": data.get('done', False),
            "task_status": data.get('task_status', 'pending'),
            "created_at": data.get('created_at', datetime.now().isoformat())
        }).execute()

        return jsonify({"success": True, "id": new_id})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/task/<task_id>', methods=['PUT'])
def update_task(task_id):
    """更新任务"""
    try:
        sb = get_supabase()
        data = request.json
        update_data = {}
        for field in ['title', 'project_id', 'center_id', 'priority', 'ability_type',
                       'due_date', 'done', 'task_status']:
            if field in data:
                update_data[field] = data[field]

        sb.table('tasks').update(update_data).eq('id', task_id).execute()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/task/<task_id>', methods=['DELETE'])
def delete_task(task_id):
    """删除任务"""
    try:
        sb = get_supabase()
        sb.table('tasks').delete().eq('id', task_id).execute()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# ========== 监查问题 API ==========

@app.route('/api/findings')
def get_findings():
    """获取监查问题列表"""
    try:
        sb = get_supabase()
        project_id = request.args.get('project_id', '')

        query = sb.table('findings').select('*')
        if project_id:
            query = query.eq('project_id', project_id)

        result = query.order('created_at', desc=False).execute()
        findings = []
        for row in result.data:
            # 获取中心和项目名称
            center_name = ""
            if row.get('center_id'):
                c_result = sb.table('centers').select('name', 'code').eq('id', row['center_id']).execute()
                if c_result.data:
                    c = c_result.data[0]
                    center_name = f"{c.get('code', '')} {c.get('name', '')}".strip()

            p_result = sb.table('projects').select('name').eq('id', row['project_id']).execute()
            project_name = p_result.data[0]['name'] if p_result.data else ""

            findings.append({
                "id": row['id'],
                "project_id": row['project_id'],
                "project_name": project_name,
                "center_id": row.get('center_id', ''),
                "center_name": center_name,
                "category": row.get('category', ''),
                "description": row['description'],
                "severity": row.get('severity', 'Minor'),
                "status": row.get('status', 'Open'),
                "found_date": row.get('found_date', ''),
                "due_date": row.get('due_date', ''),
                "corrective_action": row.get('corrective_action', ''),
                "finding_number": row.get('finding_number', ''),
                "created_at": row.get('created_at', ''),
                "updated_at": row.get('updated_at', '')
            })
        return jsonify({"success": True, "findings": findings})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/finding', methods=['POST'])
def create_finding():
    """创建监查问题"""
    try:
        sb = get_supabase()
        data = request.json
        new_id = gen_id()

        sb.table('findings').insert({
            "id": new_id,
            "project_id": data.get('project_id', ''),
            "center_id": data.get('center_id', ''),
            "category": data.get('category', ''),
            "description": data.get('description', ''),
            "severity": data.get('severity', 'Minor'),
            "status": data.get('status', 'Open'),
            "found_date": data.get('found_date', ''),
            "due_date": data.get('due_date', ''),
            "corrective_action": data.get('corrective_action', ''),
            "finding_number": data.get('finding_number', ''),
            "created_at": datetime.now().isoformat(),
            "updated_at": datetime.now().isoformat()
        }).execute()

        return jsonify({"success": True, "id": new_id})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/finding/<finding_id>', methods=['PUT'])
def update_finding(finding_id):
    """更新监查问题"""
    try:
        sb = get_supabase()
        data = request.json
        update_data = {"updated_at": datetime.now().isoformat()}
        for field in ['project_id', 'center_id', 'category', 'description', 'severity',
                       'status', 'found_date', 'due_date', 'corrective_action', 'finding_number']:
            if field in data:
                update_data[field] = data[field]

        sb.table('findings').update(update_data).eq('id', finding_id).execute()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/finding/<finding_id>', methods=['DELETE'])
def delete_finding(finding_id):
    """删除监查问题"""
    try:
        sb = get_supabase()
        sb.table('findings').delete().eq('id', finding_id).execute()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# ========== 伦理递交 API ==========

@app.route('/api/center/<center_id>/ethics')
def get_ethics(center_id):
    """获取伦理递交列表"""
    try:
        sb = get_supabase()
        result = sb.table('ethics_submissions').select('*').eq('center_id', center_id).order('created_at', desc=False).execute()
        return jsonify({"success": True, "ethics": result.data})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/ethics', methods=['POST'])
def create_ethics():
    """创建伦理递交"""
    try:
        sb = get_supabase()
        data = request.json
        new_id = gen_id()

        sb.table('ethics_submissions').insert({
            "id": new_id,
            "center_id": data.get('center_id', ''),
            "project_id": data.get('project_id', ''),
            "item_name": data.get('item_name', ''),
            "submission_date": data.get('submission_date', ''),
            "approval_date": data.get('approval_date', ''),
            "status": data.get('status', 'pending'),
            "notes": data.get('notes', ''),
            "created_at": datetime.now().isoformat()
        }).execute()

        return jsonify({"success": True, "id": new_id})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/ethics/<ethics_id>', methods=['PUT'])
def update_ethics(ethics_id):
    """更新伦理递交"""
    try:
        sb = get_supabase()
        data = request.json
        sb.table('ethics_submissions').update(data).eq('id', ethics_id).execute()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/ethics/<ethics_id>', methods=['DELETE'])
def delete_ethics(ethics_id):
    """删除伦理递交"""
    try:
        sb = get_supabase()
        sb.table('ethics_submissions').delete().eq('id', ethics_id).execute()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# ========== 方案偏离 API ==========

@app.route('/api/center/<center_id>/deviations')
def get_deviations(center_id):
    """获取方案偏离列表"""
    try:
        sb = get_supabase()
        result = sb.table('protocol_deviations').select('*').eq('center_id', center_id).order('created_at', desc=False).execute()
        return jsonify({"success": True, "deviations": result.data})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/deviation', methods=['POST'])
def create_deviation():
    """创建方案偏离"""
    try:
        sb = get_supabase()
        data = request.json
        new_id = gen_id()

        sb.table('protocol_deviations').insert({
            "id": new_id,
            "center_id": data.get('center_id', ''),
            "project_id": data.get('project_id', ''),
            "subject_id": data.get('subject_id', ''),
            "description": data.get('description', ''),
            "deviation_date": data.get('deviation_date', ''),
            "reported_date": data.get('reported_date', ''),
            "status": data.get('status', 'Open'),
            "notes": data.get('notes', ''),
            "created_at": datetime.now().isoformat()
        }).execute()

        return jsonify({"success": True, "id": new_id})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/deviation/<deviation_id>', methods=['PUT'])
def update_deviation(deviation_id):
    """更新方案偏离"""
    try:
        sb = get_supabase()
        data = request.json
        sb.table('protocol_deviations').update(data).eq('id', deviation_id).execute()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/deviation/<deviation_id>', methods=['DELETE'])
def delete_deviation(deviation_id):
    """删除方案偏离"""
    try:
        sb = get_supabase()
        sb.table('protocol_deviations').delete().eq('id', deviation_id).execute()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# ========== 研究人员 API ==========

@app.route('/api/center/<center_id>/staff')
def get_staff(center_id):
    """获取研究人员列表"""
    try:
        sb = get_supabase()
        result = sb.table('research_staff').select('*').eq('center_id', center_id).order('created_at', desc=False).execute()
        return jsonify({"success": True, "staff": result.data})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/staff', methods=['POST'])
def create_staff():
    """创建研究人员"""
    try:
        sb = get_supabase()
        data = request.json
        new_id = gen_id()

        sb.table('research_staff').insert({
            "id": new_id,
            "center_id": data.get('center_id', ''),
            "project_id": data.get('project_id', ''),
            "name": data.get('name', ''),
            "role": data.get('role', ''),
            "gcp_cert": data.get('gcp_cert', ''),
            "gcp_expiry": data.get('gcp_expiry', ''),
            "cv_date": data.get('cv_date', ''),
            "notes": data.get('notes', ''),
            "created_at": datetime.now().isoformat()
        }).execute()

        return jsonify({"success": True, "id": new_id})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/staff/<staff_id>', methods=['PUT'])
def update_staff(staff_id):
    """更新研究人员"""
    try:
        sb = get_supabase()
        data = request.json
        sb.table('research_staff').update(data).eq('id', staff_id).execute()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/staff/<staff_id>', methods=['DELETE'])
def delete_staff(staff_id):
    """删除研究人员"""
    try:
        sb = get_supabase()
        sb.table('research_staff').delete().eq('id', staff_id).execute()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# ========== 统计 API ==========

@app.route('/api/stats')
def get_stats():
    """获取统计数据"""
    try:
        sb = get_supabase()

        total_projects = sb.table('projects').select('id', count='exact').execute().count
        total_tasks = sb.table('tasks').select('id', count='exact').execute().count
        done_tasks = sb.table('tasks').select('id', count='exact').eq('done', True).execute().count
        pending_tasks = total_tasks - done_tasks

        total_findings = sb.table('findings').select('id', count='exact').execute().count
        open_findings = sb.table('findings').select('id', count='exact').eq('status', 'Open').execute().count

        # 计算逾期任务
        today_str = date.today().isoformat()
        all_tasks = sb.table('tasks').select('due_date', 'done').eq('done', False).execute().data
        overdue = sum(1 for t in all_tasks if t.get('due_date') and t['due_date'] < today_str)

        # 计算即将到期（7天内）
        from datetime import timedelta
        soon_str = (date.today() + timedelta(days=7)).isoformat()
        due_soon = sum(1 for t in all_tasks if t.get('due_date') and today_str <= t['due_date'] <= soon_str)

        # 高优先级
        high_priority = sb.table('tasks').select('id', count='exact').eq('priority', 'high').eq('done', False).execute().count

        # 等待CRC
        waiting_crc = sb.table('tasks').select('id', count='exact').eq('task_status', 'waiting_crc').execute().count

        # 活跃项目（有未完成任务的项目）
        active_projects = 0
        projects = sb.table('projects').select('id').execute().data
        for p in projects:
            p_pending = sb.table('tasks').select('id', count='exact').eq('project_id', p['id']).eq('done', False).execute().count
            if p_pending > 0:
                active_projects += 1

        # 中心进度
        center_progress = []
        centers = sb.table('centers').select('id', 'name', 'code', 'project_id').execute().data
        for c in centers:
            c_total = sb.table('tasks').select('id', count='exact').eq('center_id', c['id']).execute().count
            c_done = sb.table('tasks').select('id', count='exact').eq('center_id', c['id']).eq('done', True).execute().count
            pct = int(c_done / c_total * 100) if c_total > 0 else 0
            center_progress.append({
                "id": c['id'],
                "code": c.get('code', ''),
                "name": c['name'],
                "total": c_total,
                "done": c_done,
                "pct": pct
            })

        return jsonify({
            "success": True,
            "stats": {
                "total_projects": total_projects,
                "active_projects": active_projects,
                "total_tasks": total_tasks,
                "done_tasks": done_tasks,
                "pending_tasks": pending_tasks,
                "overdue_tasks": overdue,
                "due_soon": due_soon,
                "high_priority": high_priority,
                "waiting_crc_tasks": waiting_crc,
                "total_findings": total_findings,
                "open_findings": open_findings,
                "center_progress": center_progress
            }
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# ========== 导出 Excel API ==========

@app.route('/api/export/<export_type>')
def export_excel(export_type):
    """导出 Excel 文件"""
    try:
        import openpyxl
        wb = openpyxl.Workbook()

        sb = get_supabase()

        if export_type == 'tasks' or export_type == 'all':
            ws = wb.active
            ws.title = "待办事项"
            ws.append(["ID", "标题", "项目", "中心", "优先级", "能力类型", "截止日期", "状态", "创建时间"])
            tasks_result = sb.table('tasks').select('*').execute()
            for t in tasks_result.data:
                ws.append([
                    t['id'], t['title'], t.get('project_id', ''), t.get('center_id', ''),
                    t.get('priority', ''), t.get('ability_type', ''), t.get('due_date', ''),
                    "已完成" if t.get('done') else "待办", t.get('created_at', '')
                ])

        if export_type == 'findings' or export_type == 'all':
            ws = wb.create_sheet("监查问题")
            ws.append(["ID", "项目", "中心", "分类", "描述", "严重程度", "状态", "发现日期", "截止日期", "纠正措施"])
            findings_result = sb.table('findings').select('*').execute()
            for f in findings_result.data:
                ws.append([
                    f['id'], f.get('project_id', ''), f.get('center_id', ''), f.get('category', ''),
                    f['description'], f.get('severity', ''), f.get('status', ''),
                    f.get('found_date', ''), f.get('due_date', ''), f.get('corrective_action', '')
                ])

        if export_type == 'centers' or export_type == 'all':
            ws = wb.create_sheet("中心信息")
            ws.append(["ID", "项目ID", "编号", "名称", "PI", "PI电话", "CRC", "CRC电话", "地址", "备注"])
            centers_result = sb.table('centers').select('*').execute()
            for c in centers_result.data:
                ws.append([
                    c['id'], c['project_id'], c.get('code', ''), c['name'],
                    c.get('pi_name', ''), c.get('pi_phone', ''),
                    c.get('contact_crc', ''), c.get('contact_crc_phone', ''),
                    c.get('address', ''), c.get('notes', '')
                ])

        if export_type == 'projects' or export_type == 'all':
            ws = wb.create_sheet("项目信息")
            ws.append(["ID", "名称", "编号", "阶段", "DBL日期", "备注", "创建时间"])
            projects_result = sb.table('projects').select('*').execute()
            for p in projects_result.data:
                ws.append([
                    p['id'], p['name'], p.get('code', ''), p.get('stage', ''),
                    p.get('dbl_date', ''), p.get('notes', ''), p.get('created_at', '')
                ])

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"cra-portal-{export_type}-{datetime.now().strftime('%Y%m%d')}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# ========== 数据备份 API ==========

@app.route('/api/backup')
def backup_data():
    """导出 JSON 数据备份"""
    try:
        sb = get_supabase()
        backup = {
            "backup_time": datetime.now().isoformat(),
            "data": {}
        }

        tables = ['projects', 'centers', 'tasks', 'findings',
                  'ethics_submissions', 'protocol_deviations', 'research_staff', 'startup_tasks']

        for table_name in tables:
            result = sb.table(table_name).select('*').execute()
            backup["data"][table_name] = result.data

        json_str = json.dumps(backup, ensure_ascii=False, indent=2, default=str)
        filename = f"cra-portal-backup-{datetime.now().strftime('%Y%m%d')}.json"

        return Response(
            json_str,
            mimetype='application/json',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# ========== 启动任务 API ==========

@app.route('/api/startup-tasks')
def get_startup_tasks():
    """获取启动任务列表"""
    try:
        sb = get_supabase()
        result = sb.table('startup_tasks').select('*').order('created_at', desc=False).execute()
        return jsonify({"success": True, "tasks": result.data})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/startup-task', methods=['POST'])
def create_startup_task():
    """创建启动任务"""
    try:
        sb = get_supabase()
        data = request.json
        new_id = gen_id()

        sb.table('startup_tasks').insert({
            "id": new_id,
            "name": data.get('name', ''),
            "description": data.get('description', ''),
            "done": False,
            "created_at": datetime.now().isoformat()
        }).execute()

        return jsonify({"success": True, "id": new_id})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/startup-task/<task_id>', methods=['PUT'])
def update_startup_task(task_id):
    """更新启动任务"""
    try:
        sb = get_supabase()
        data = request.json
        sb.table('startup_tasks').update(data).eq('id', task_id).execute()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/startup-task/<task_id>', methods=['DELETE'])
def delete_startup_task(task_id):
    """删除启动任务"""
    try:
        sb = get_supabase()
        sb.table('startup_tasks').delete().eq('id', task_id).execute()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# ========== 健康检查 ==========

@app.route('/api/health')
def health_check():
    """健康检查"""
    try:
        sb = get_supabase()
        result = sb.table('projects').select('id', count='exact').limit(1).execute()
        return jsonify({"success": True, "status": "healthy", "database": "supabase", "project_count": result.count})
    except Exception as e:
        return jsonify({"success": False, "status": "unhealthy", "error": str(e)}), 500


# ========== 主程序 ==========

if __name__ == '__main__':
    print("=" * 50)
    print("CRA Portal v2.0 (Supabase) 启动中...")
    print(f"Supabase URL: {SUPABASE_URL or '未配置'}")
    print("访问地址: http://localhost:8080")
    print("=" * 50)
    port = int(os.environ.get('PORT', 8080))
    app.run(host='0.0.0.0', port=port, debug=True)
